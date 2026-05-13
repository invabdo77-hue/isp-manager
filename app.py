from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
import sqlite3
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'isp_manager_secret_key_2024')

DATABASE_URL = os.environ.get('DATABASE_URL')
USE_POSTGRES = bool(DATABASE_URL)

def get_db():
    if USE_POSTGRES:
        import psycopg2
        import psycopg2.extras
        url = DATABASE_URL
        if 'sslmode' not in url.lower():
            separator = '&' if '?' in url else '?'
            url = f"{url}{separator}sslmode=require"
        conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
        return conn
    else:
        DATABASE = os.path.join(os.path.dirname(__file__), 'isp_manager.db')
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        return conn

def get_placeholder():
    return '%s' if USE_POSTGRES else '?'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        cur.execute(f'SELECT * FROM users WHERE id = {ph}', (user_id,))
        user = cur.fetchone()
        if user:
            return User(user['id'], user['username'], user['role'])
        return None
    except Exception:
        return None
    finally:
        if conn:
            conn.close()

def require_admin(f):
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Acceso solo para administradores', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated.__name__ = f.__name__
    return decorated

@app.route('/setup')
def setup():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
    except Exception as e:
        return f"Error conectando a la base de datos: {e}", 500

    try:
        if USE_POSTGRES:
            cur.execute('''CREATE TABLE IF NOT EXISTS plans (
                id SERIAL PRIMARY KEY, name TEXT NOT NULL, speed TEXT NOT NULL,
                price REAL NOT NULL, description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, role TEXT DEFAULT 'technician',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS clients (
                id SERIAL PRIMARY KEY, first_name TEXT NOT NULL, last_name TEXT NOT NULL,
                cedula TEXT, phone TEXT, email TEXT, address TEXT, router_model TEXT,
                router_serial TEXT, router_mac TEXT, ip_address TEXT, nap_number TEXT,
                potencia TEXT, plan_id INTEGER, connection_status TEXT DEFAULT 'active',
                registration_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS monthly_payments (
                id SERIAL PRIMARY KEY, client_id INTEGER NOT NULL, amount REAL NOT NULL,
                month TEXT NOT NULL, payment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                method TEXT DEFAULT 'efectivo', status TEXT DEFAULT 'paid', notes TEXT)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS other_incomes (
                id SERIAL PRIMARY KEY, description TEXT NOT NULL, amount REAL NOT NULL,
                category TEXT DEFAULT 'otro', income_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP, notes TEXT)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS expenses (
                id SERIAL PRIMARY KEY, description TEXT NOT NULL, amount REAL NOT NULL,
                category TEXT DEFAULT 'otro', expense_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP, notes TEXT)''')
        else:
            cur.execute('''CREATE TABLE IF NOT EXISTS plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, speed TEXT NOT NULL,
                price REAL NOT NULL, description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL, role TEXT DEFAULT 'technician',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT, first_name TEXT NOT NULL, last_name TEXT NOT NULL,
                cedula TEXT, phone TEXT, email TEXT, address TEXT, router_model TEXT,
                router_serial TEXT, router_mac TEXT, ip_address TEXT, nap_number TEXT,
                potencia TEXT, plan_id INTEGER, connection_status TEXT DEFAULT 'active',
                registration_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS monthly_payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT, client_id INTEGER NOT NULL, amount REAL NOT NULL,
                month TEXT NOT NULL, payment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                method TEXT DEFAULT 'efectivo', status TEXT DEFAULT 'paid', notes TEXT)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS other_incomes (
                id INTEGER PRIMARY KEY AUTOINCREMENT, description TEXT NOT NULL, amount REAL NOT NULL,
                category TEXT DEFAULT 'otro', income_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP, notes TEXT)''')
            cur.execute('''CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT, description TEXT NOT NULL, amount REAL NOT NULL,
                category TEXT DEFAULT 'otro', expense_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP, notes TEXT)''')

        cur.execute('SELECT COUNT(*) as count FROM users')
        if cur.fetchone()['count'] == 0:
            cur.execute(f"INSERT INTO users (username, password, role) VALUES ({ph}, {ph}, {ph})",
                       ('admin', generate_password_hash('admin123'), 'admin'))
            cur.execute(f"INSERT INTO users (username, password, role) VALUES ({ph}, {ph}, {ph})",
                       ('tecnico1', generate_password_hash('tecnico123'), 'technician'))

        cur.execute('SELECT COUNT(*) as count FROM plans')
        if cur.fetchone()['count'] == 0:
            cur.execute(f"INSERT INTO plans (name, speed, price, description) VALUES ({ph}, {ph}, {ph}, {ph})",
                       ('100 Mbps', '100 Mbps', 25.00, 'Plan 100 Mbps'))
            cur.execute(f"INSERT INTO plans (name, speed, price, description) VALUES ({ph}, {ph}, {ph}, {ph})",
                       ('200 Mbps', '200 Mbps', 30.00, 'Plan 200 Mbps'))
            cur.execute(f"INSERT INTO plans (name, speed, price, description) VALUES ({ph}, {ph}, {ph}, {ph})",
                       ('300 Mbps', '300 Mbps', 40.00, 'Plan 300 Mbps'))

        conn.commit()
        return f"Base de datos inicializada correctamente! (PostgreSQL: {USE_POSTGRES})", 200
    except Exception as e:
        if conn:
            conn.rollback()
        return f"Error inicializando la base de datos: {e}", 500
    finally:
        if conn:
            conn.close()



@app.route('/reset-users')
def reset_users():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        cur.execute("DELETE FROM users WHERE username IN ('admin', 'tecnico1')")
        cur.execute(f"INSERT INTO users (username, password, role) VALUES ({ph}, {ph}, {ph})",
                   ('admin', generate_password_hash('admin123'), 'admin'))
        cur.execute(f"INSERT INTO users (username, password, role) VALUES ({ph}, {ph}, {ph})",
                   ('tecnico1', generate_password_hash('tecnico123'), 'technician'))
        conn.commit()
        return "Usuarios reiniciados! admin/admin123 y tecnico1/tecnico123", 200
    except Exception as e:
        if conn:
            conn.rollback()
        return f"Error reiniciando usuarios: {e}", 500
    finally:
        if conn:
            conn.close()

@app.route('/debug-login')
def debug_login():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT id, username, role, password FROM users')
        users = cur.fetchall()
        html = '<h1>Usuarios en DB:</h1><ul>'
        for u in users:
            html += f"<li>ID: {u['id']}, User: {u['username']}, Role: {u['role']}<br>Hash: {u['password'][:50]}...</li>"
        html += '</ul><p>Prueba: admin / admin123</p>'
        return html
    except Exception as e:
        return f"Error consultando usuarios: {e}", 500
    finally:
        if conn:
            conn.close()

@app.route('/')
@login_required
def index():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) as count FROM clients')
        total_clients = cur.fetchone()['count']
        cur.execute("SELECT COUNT(*) as count FROM clients WHERE connection_status = 'active'")
        active_clients = cur.fetchone()['count']
        cur.execute('''SELECT c.*, p.name as plan_name, p.price as plan_price
            FROM clients c LEFT JOIN plans p ON c.plan_id = p.id
            ORDER BY c.registration_date DESC LIMIT 5''')
        recent_clients = cur.fetchall()
        return render_template('index.html', total_clients=total_clients,
                               active_clients=active_clients, recent_clients=recent_clients)
    except Exception as e:
        flash(f'Error cargando datos: {e}', 'danger')
        return render_template('index.html', total_clients=0, active_clients=0, recent_clients=[])
    finally:
        if conn:
            conn.close()

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            flash('Usuario y contraseña son requeridos', 'danger')
            return render_template('login.html')
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f"SELECT * FROM users WHERE username = {ph}", (username,))
            user = cur.fetchone()
        except Exception as e:
            flash(f'Error de base de datos: {e}', 'danger')
            return render_template('login.html')
        finally:
            if conn:
                conn.close()
        if user and check_password_hash(user['password'], password):
            login_user(User(user['id'], user['username'], user['role']))
            return redirect(url_for('index'))
        flash('Usuario o contraseña incorrectos', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/clients')
@login_required
def clients_list():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''SELECT c.*, p.name as plan_name, p.speed as plan_speed, p.price as plan_price
            FROM clients c LEFT JOIN plans p ON c.plan_id = p.id
            ORDER BY c.first_name, c.last_name''')
        clients = cur.fetchall()
        return render_template('clients.html', clients=clients)
    except Exception as e:
        flash(f'Error cargando clientes: {e}', 'danger')
        return render_template('clients.html', clients=[])
    finally:
        if conn:
            conn.close()

@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def client_new():
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        if not first_name or not last_name:
            flash('Nombre y apellido son requeridos', 'danger')
            conn = None
            try:
                conn = get_db()
                cur = conn.cursor()
                cur.execute('SELECT * FROM plans ORDER BY price')
                plans = cur.fetchall()
            finally:
                if conn:
                    conn.close()
            return render_template('client_form.html', plans=plans, client=None)

        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            vals = [first_name, last_name]
            cols = ['first_name', 'last_name']

            for field in ['cedula', 'phone', 'email', 'address', 'router_model', 'router_serial',
                          'router_mac', 'ip_address', 'nap_number', 'potencia']:
                val = request.form.get(field, '').strip()
                if val:
                    vals.append(val)
                    cols.append(field)

            plan_id_str = request.form.get('plan_id', '').strip()
            if plan_id_str:
                try:
                    vals.append(int(plan_id_str))
                    cols.append('plan_id')
                except ValueError:
                    pass

            vals.append(request.form.get('connection_status', 'active') if current_user.role == 'admin' else 'active')
            cols.append('connection_status')

            ph = get_placeholder()
            placeholders = ','.join([ph] * len(cols))
            sql = f'INSERT INTO clients ({", ".join(cols)}) VALUES ({placeholders})'
            cur.execute(sql, vals)
            conn.commit()
            flash('Cliente creado exitosamente', 'success')
            return redirect(url_for('clients_list'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error creando cliente: {e}', 'danger')
        finally:
            if conn:
                conn.close()

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT * FROM plans ORDER BY price')
        plans = cur.fetchall()
        return render_template('client_form.html', plans=plans, client=None)
    except Exception as e:
        flash(f'Error cargando formulario: {e}', 'danger')
        return render_template('client_form.html', plans=[], client=None)
    finally:
        if conn:
            conn.close()

@app.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    if request.method == 'POST':
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            updates = []
            values = []
            for field in ['first_name', 'last_name', 'cedula', 'phone', 'email', 'address',
                          'router_model', 'router_serial', 'router_mac', 'ip_address',
                          'nap_number', 'potencia', 'plan_id', 'connection_status']:
                val = request.form.get(field, '').strip()
                if val:
                    updates.append(f'{field}={get_placeholder()}')
                    if field == 'plan_id':
                        try:
                            val = int(val)
                        except ValueError:
                            continue
                    values.append(val)
            if updates:
                values.append(id)
                cur.execute(f'UPDATE clients SET {", ".join(updates)} WHERE id={get_placeholder()}', values)
                conn.commit()
            flash('Cliente actualizado', 'success')
            return redirect(url_for('clients_list'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error actualizando cliente: {e}', 'danger')
            return redirect(url_for('clients_list'))
        finally:
            if conn:
                conn.close()

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f'SELECT * FROM clients WHERE id = {get_placeholder()}', (id,))
        client = cur.fetchone()
        if not client:
            flash('Cliente no encontrado', 'danger')
            return redirect(url_for('clients_list'))
        cur.execute('SELECT * FROM plans ORDER BY price')
        plans = cur.fetchall()
        return render_template('client_form.html', plans=plans, client=client)
    except Exception as e:
        flash(f'Error cargando cliente: {e}', 'danger')
        return redirect(url_for('clients_list'))
    finally:
        if conn:
            conn.close()

@app.route('/clients/<int:id>/delete', methods=['POST'])
@login_required
@require_admin
def client_delete(id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f'DELETE FROM clients WHERE id = {get_placeholder()}', (id,))
        conn.commit()
        flash('Cliente eliminado', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error eliminando cliente: {e}', 'danger')
    finally:
        if conn:
            conn.close()
    return redirect(url_for('clients_list'))

@app.route('/plans')
@login_required
@require_admin
def plans_list():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''SELECT p.*, COUNT(c.id) as client_count
            FROM plans p LEFT JOIN clients c ON p.id = c.plan_id
            GROUP BY p.id ORDER BY p.price''')
        plans = cur.fetchall()
        return render_template('plans.html', plans=plans)
    except Exception as e:
        flash(f'Error cargando planes: {e}', 'danger')
        return render_template('plans.html', plans=[])
    finally:
        if conn:
            conn.close()

@app.route('/plans/new', methods=['GET', 'POST'])
@login_required
@require_admin
def plan_new():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        speed = request.form.get('speed', '').strip()
        price_str = request.form.get('price', '').strip()
        if not name or not speed or not price_str:
            flash('Nombre, velocidad y precio son requeridos', 'danger')
            return render_template('plan_form.html', plan=None)
        try:
            price = float(price_str)
        except ValueError:
            flash('El precio debe ser un número válido', 'danger')
            return render_template('plan_form.html', plan=None)
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f'INSERT INTO plans (name, speed, price, description) VALUES ({ph}, {ph}, {ph}, {ph})',
                       (name, speed, price, request.form.get('description', '').strip()))
            conn.commit()
            flash('Plan creado', 'success')
            return redirect(url_for('plans_list'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error creando plan: {e}', 'danger')
        finally:
            if conn:
                conn.close()
    return render_template('plan_form.html', plan=None)

@app.route('/plans/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@require_admin
def plan_edit(id):
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        speed = request.form.get('speed', '').strip()
        price_str = request.form.get('price', '').strip()
        if not name or not speed or not price_str:
            flash('Nombre, velocidad y precio son requeridos', 'danger')
            return redirect(url_for('plan_edit', id=id))
        try:
            price = float(price_str)
        except ValueError:
            flash('El precio debe ser un número válido', 'danger')
            return redirect(url_for('plan_edit', id=id))
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f'UPDATE plans SET name={ph}, speed={ph}, price={ph}, description={ph} WHERE id={ph}',
                       (name, speed, price, request.form.get('description', '').strip(), id))
            conn.commit()
            flash('Plan actualizado', 'success')
            return redirect(url_for('plans_list'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error actualizando plan: {e}', 'danger')
            return redirect(url_for('plans_list'))
        finally:
            if conn:
                conn.close()

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        cur.execute(f'SELECT * FROM plans WHERE id = {ph}', (id,))
        plan = cur.fetchone()
        if not plan:
            flash('Plan no encontrado', 'danger')
            return redirect(url_for('plans_list'))
        return render_template('plan_form.html', plan=plan)
    except Exception as e:
        flash(f'Error cargando plan: {e}', 'danger')
        return redirect(url_for('plans_list'))
    finally:
        if conn:
            conn.close()

@app.route('/plans/<int:id>/delete', methods=['POST'])
@login_required
@require_admin
def plan_delete(id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f'DELETE FROM plans WHERE id = {get_placeholder()}', (id,))
        conn.commit()
        flash('Plan eliminado', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error eliminando plan: {e}', 'danger')
    finally:
        if conn:
            conn.close()
    return redirect(url_for('plans_list'))

@app.route('/export/plans')
@login_required
def export_plans():
    from openpyxl import Workbook
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT * FROM plans ORDER BY price')
        plans = cur.fetchall()
    finally:
        if conn:
            conn.close()
    wb = Workbook()
    ws = wb.active
    ws.append(['ID', 'Nombre', 'Velocidad', 'Precio', 'Descripcion'])
    for p in plans:
        ws.append([p['id'], p['name'], p['speed'], p['price'], p['description']])
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    from flask import make_response
    return make_response(output.getvalue(), {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': 'attachment; filename=planes.xlsx'})

@app.route('/finances')
@login_required
@require_admin
def finances():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT COALESCE(SUM(amount), 0) as total FROM monthly_payments WHERE status = 'paid'")
        total_monthly = cur.fetchone()['total']
        cur.execute("SELECT COALESCE(SUM(amount), 0) as total FROM other_incomes")
        total_other_incomes = cur.fetchone()['total']
        cur.execute("SELECT COALESCE(SUM(amount), 0) as total FROM expenses")
        total_expenses = cur.fetchone()['total']
        balance = total_monthly + total_other_incomes - total_expenses

        cur.execute('''SELECT mp.*, c.first_name, c.last_name, p.name as plan_name
            FROM monthly_payments mp JOIN clients c ON mp.client_id = c.id
            LEFT JOIN plans p ON c.plan_id = p.id ORDER BY mp.payment_date DESC''')
        monthly_payments = cur.fetchall()

        cur.execute('SELECT * FROM other_incomes ORDER BY income_date DESC')
        other_incomes = cur.fetchall()
        cur.execute('SELECT * FROM expenses ORDER BY expense_date DESC')
        expenses = cur.fetchall()
        cur.execute('SELECT category, SUM(amount) as total FROM other_incomes GROUP BY category')
        income_by_category = cur.fetchall()
        cur.execute('SELECT category, SUM(amount) as total FROM expenses GROUP BY category')
        expense_by_category = cur.fetchall()

        return render_template('finances.html', total_monthly=total_monthly,
                               total_other_incomes=total_other_incomes,
                               total_expenses=total_expenses, balance=balance,
                               monthly_payments=monthly_payments, other_incomes=other_incomes,
                               expenses=expenses, income_by_category=income_by_category,
                               expense_by_category=expense_by_category)
    except Exception as e:
        flash(f'Error cargando finanzas: {e}', 'danger')
        return render_template('finances.html', total_monthly=0, total_other_incomes=0,
                               total_expenses=0, balance=0, monthly_payments=[],
                               other_incomes=[], expenses=[], income_by_category=[],
                               expense_by_category=[])
    finally:
        if conn:
            conn.close()

@app.route('/finances/payment/new', methods=['GET', 'POST'])
@login_required
@require_admin
def payment_new():
    if request.method == 'POST':
        client_id_str = request.form.get('client_id', '').strip()
        amount_str = request.form.get('amount', '').strip()
        month = request.form.get('month', '').strip()
        if not client_id_str or not amount_str or not month:
            flash('Cliente, monto y mes son requeridos', 'danger')
            return redirect(url_for('payment_new'))
        try:
            client_id = int(client_id_str)
            amount = float(amount_str)
        except ValueError:
            flash('Cliente y monto deben ser valores válidos', 'danger')
            return redirect(url_for('payment_new'))
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f'INSERT INTO monthly_payments (client_id, amount, month, payment_date, method, status, notes) VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})',
                       (client_id, amount, month,
                        request.form.get('payment_date', datetime.now().strftime('%Y-%m-%d')),
                        request.form.get('method', 'efectivo'), 'paid', request.form.get('notes')))
            cur.execute(f"UPDATE clients SET connection_status = 'active' WHERE id = {ph} AND connection_status = 'cut'", (client_id,))
            conn.commit()
            flash('Pago registrado', 'success')
            return redirect(url_for('finances'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error registrando pago: {e}', 'danger')
            return redirect(url_for('payment_new'))
        finally:
            if conn:
                conn.close()

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT id, first_name, last_name FROM clients ORDER BY first_name, last_name')
        clients = cur.fetchall()
        return render_template('payment_form.html', clients=clients,
                               current_month=datetime.now().strftime('%Y-%m'), payment=None)
    except Exception as e:
        flash(f'Error cargando formulario: {e}', 'danger')
        return redirect(url_for('finances'))
    finally:
        if conn:
            conn.close()

@app.route('/finances/income/new', methods=['GET', 'POST'])
@login_required
@require_admin
def income_new():
    if request.method == 'POST':
        description = request.form.get('description', '').strip()
        amount_str = request.form.get('amount', '').strip()
        if not description or not amount_str:
            flash('Descripción y monto son requeridos', 'danger')
            return render_template('income_form.html', income=None)
        try:
            amount = float(amount_str)
        except ValueError:
            flash('El monto debe ser un número válido', 'danger')
            return render_template('income_form.html', income=None)
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f'INSERT INTO other_incomes (description, amount, category, income_date, notes) VALUES ({ph}, {ph}, {ph}, {ph}, {ph})',
                       (description, amount, request.form.get('category', 'otro'),
                        request.form.get('income_date', datetime.now().strftime('%Y-%m-%d')),
                        request.form.get('notes')))
            conn.commit()
            flash('Ingreso registrado', 'success')
            return redirect(url_for('finances'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error registrando ingreso: {e}', 'danger')
        finally:
            if conn:
                conn.close()
    return render_template('income_form.html', income=None)

@app.route('/finances/expense/new', methods=['GET', 'POST'])
@login_required
@require_admin
def expense_new():
    if request.method == 'POST':
        description = request.form.get('description', '').strip()
        amount_str = request.form.get('amount', '').strip()
        if not description or not amount_str:
            flash('Descripción y monto son requeridos', 'danger')
            return render_template('expense_form.html', expense=None)
        try:
            amount = float(amount_str)
        except ValueError:
            flash('El monto debe ser un número válido', 'danger')
            return render_template('expense_form.html', expense=None)
        conn = None
        try:
            conn = get_db()
            cur = conn.cursor()
            ph = get_placeholder()
            cur.execute(f'INSERT INTO expenses (description, amount, category, expense_date, notes) VALUES ({ph}, {ph}, {ph}, {ph}, {ph})',
                       (description, amount, request.form.get('category', 'otro'),
                        request.form.get('expense_date', datetime.now().strftime('%Y-%m-%d')),
                        request.form.get('notes')))
            conn.commit()
            flash('Gasto registrado', 'success')
            return redirect(url_for('finances'))
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error registrando gasto: {e}', 'danger')
        finally:
            if conn:
                conn.close()
    return render_template('expense_form.html', expense=None)

@app.route('/finances/client/<int:id>/payments')
@login_required
@require_admin
def client_payments(id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        cur.execute(f'SELECT * FROM clients WHERE id = {ph}', (id,))
        client = cur.fetchone()
        if not client:
            flash('Cliente no encontrado', 'danger')
            return redirect(url_for('finances'))
        cur.execute(f'SELECT * FROM monthly_payments WHERE client_id = {ph} ORDER BY payment_date DESC', (id,))
        payments = cur.fetchall()
        return render_template('client_payments.html', client=client, payments=payments)
    except Exception as e:
        flash(f'Error cargando pagos: {e}', 'danger')
        return redirect(url_for('finances'))
    finally:
        if conn:
            conn.close()

@app.route('/api/clients/search')
@login_required
def api_client_search():
    query = request.args.get('q', '').strip()
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        cur.execute(f'SELECT id, first_name, last_name, cedula FROM clients WHERE first_name LIKE {ph} OR last_name LIKE {ph} OR cedula LIKE {ph} LIMIT 10',
                   (f'%{query}%', f'%{query}%', f'{query}%'))
        results = cur.fetchall()
        return jsonify([dict(c) for c in results])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/export/clients')
@login_required
def export_clients():
    from openpyxl import Workbook
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''SELECT c.*, p.name as plan_name, p.speed as plan_speed, p.price as plan_price
            FROM clients c LEFT JOIN plans p ON c.plan_id = p.id''')
        clients = cur.fetchall()
    finally:
        if conn:
            conn.close()
    wb = Workbook()
    ws = wb.active
    ws.append(['ID', 'Nombre', 'Apellido', 'Cedula', 'Telefono', 'Email', 'Direccion', 'Router', 'Serial', 'MAC', 'IP', 'NAP', 'Potencia', 'Plan', 'Velocidad', 'Precio', 'Estado', 'Fecha'])
    for c in clients:
        ws.append([c['id'], c['first_name'], c['last_name'], c['cedula'], c['phone'], c['email'], c['address'], c['router_model'], c['router_serial'], c['router_mac'], c['ip_address'], c['nap_number'], c['potencia'], c['plan_name'], c['plan_speed'], c['plan_price'], c['connection_status'], c['registration_date']])
    from io import BytesIO
    from flask import make_response
    output = BytesIO()
    wb.save(output)
    return make_response(output.getvalue(), {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': 'attachment; filename=clientes.xlsx'})

@app.route('/import/clients', methods=['POST'])
@login_required
@require_admin
def import_clients():
    from openpyxl import load_workbook
    if 'file' not in request.files:
        flash('No se selecciono archivo', 'danger')
        return redirect(url_for('clients_list'))
    file = request.files['file']
    conn = None
    try:
        wb = load_workbook(file)
        ws = wb.active
        conn = get_db()
        cur = conn.cursor()
        ph = get_placeholder()
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                try:
                    cur.execute(f'''INSERT INTO clients (first_name, last_name, cedula, phone, email, address, plan_id, connection_status)
                        VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})''',
                        (row[1], row[2], row[3], row[4], row[5], row[6], None, 'active'))
                    imported += 1
                except Exception:
                    pass
        conn.commit()
        flash(f'{imported} clientes importados', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        flash(f'Error importando clientes: {e}', 'danger')
    finally:
        if conn:
            conn.close()
    return redirect(url_for('clients_list'))

@app.route('/export/finances')
@login_required
@require_admin
def export_finances():
    from openpyxl import Workbook
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append(['ID', 'Cliente', 'Monto', 'Mes', 'Fecha', 'Metodo', 'Estado'])
        cur.execute('SELECT mp.*, c.first_name, c.last_name FROM monthly_payments mp JOIN clients c ON mp.client_id = c.id')
        for p in cur.fetchall():
            ws.append([p['id'], f"{p['first_name']} {p['last_name']}", p['amount'], p['month'], p['payment_date'], p['method'], p['status']])
        ws2 = wb.create_sheet("Ingresos")
        ws2.append(['ID', 'Descripcion', 'Monto', 'Categoria', 'Fecha'])
        cur.execute('SELECT * FROM other_incomes')
        for i in cur.fetchall():
            ws2.append([i['id'], i['description'], i['amount'], i['category'], i['income_date']])
        ws3 = wb.create_sheet("Gastos")
        ws3.append(['ID', 'Descripcion', 'Monto', 'Categoria', 'Fecha'])
        cur.execute('SELECT * FROM expenses')
        for e in cur.fetchall():
            ws3.append([e['id'], e['description'], e['amount'], e['category'], e['expense_date']])
    finally:
        if conn:
            conn.close()
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    from flask import make_response
    return make_response(output.getvalue(), {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': 'attachment; filename=finanzas.xlsx'})

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=not USE_POSTGRES, host='0.0.0.0', port=port)